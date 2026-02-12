import json
import re
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = ROOT / "数据库"
ASSETS_DIR = ROOT / "android" / "app" / "src" / "main" / "assets"

TAXONOMY_XLSX = SRC_DIR / "浮游动物分类.xlsx"
WETWEIGHT_XLSX = SRC_DIR / "评价湿重.xlsx"

OUT_TAXONOMIES = ASSETS_DIR / "taxonomies.json"
OUT_WETWEIGHTS = ASSETS_DIR / "wetweights.json"


def _normalize_lvl1(v: str) -> str:
    v = (v or "").strip()
    return {
        "轮虫": "轮虫类",
        "轮虫类": "轮虫类",
        "枝角": "枝角类",
        "枝角类": "枝角类",
        "桡足": "桡足类",
        "桡足类": "桡足类",
        "原生动物类": "原生动物",
        "原生动物": "原生动物",
    }.get(v, v)


_CN_LATIN_RE = re.compile(r"^(?P<cn>[^（(]+)[（(](?P<latin>[^）)]+)[）)]$")


def _split_cn_latin(raw: str) -> tuple[str, str | None]:
    raw = (raw or "").strip()
    if not raw:
        return "", None
    m = _CN_LATIN_RE.match(raw)
    if not m:
        return raw, None
    cn = m.group("cn").strip()
    latin = m.group("latin").strip() or None
    return cn, latin


def parse_taxonomies(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    cur: list[str | None] = [None] * 6
    entries = []
    seen = set()

    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, 7)]
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        for i, v in enumerate(row):
            if v is None:
                continue
            s = str(v).strip()
            if s != "":
                cur[i] = s

        raw_species = (cur[5] or "").strip() if cur[5] is not None else ""
        if not raw_species:
            continue

        name_cn, name_latin = _split_cn_latin(raw_species)
        if not name_cn or name_cn in seen:
            continue
        seen.add(name_cn)

        lvl1 = _normalize_lvl1(cur[0] or "")
        entry = {
            "nameCn": name_cn,
            "nameLatin": name_latin,
            "taxonomy": {
                "lvl1": lvl1,
                "lvl2": (cur[1] or "").strip(),
                "lvl3": (cur[2] or "").strip(),
                "lvl4": (cur[3] or "").strip(),
                "lvl5": (cur[4] or "").strip(),
            },
        }
        entries.append(entry)

    return {"version": 1, "entries": entries}


def parse_wetweights(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row = None
    for r in range(1, ws.max_row + 1):
        values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(str(v).strip() == "平均湿重" for v in values if v is not None):
            header_row = r
            break
    if header_row is None:
        raise RuntimeError("未找到“平均湿重”表头行")

    cur_group: str | None = None
    cur_sub: str | None = None
    entries = []
    seen = set()
    lvl1_set = {"原生动物", "轮虫", "轮虫类", "枝角", "枝角类", "桡足", "桡足类"}

    for r in range(header_row + 1, ws.max_row + 1):
        cn = ws.cell(r, 1).value
        latin = ws.cell(r, 2).value
        weight = ws.cell(r, 3).value

        if cn is None or str(cn).strip() == "":
            continue
        name_cn = str(cn).strip()
        name_latin = str(latin).strip() if latin is not None and str(latin).strip() != "" else None

        if weight is None or str(weight).strip() == "":
            # Heading row
            if name_cn in lvl1_set:
                cur_group = _normalize_lvl1(name_cn)
                cur_sub = None
            else:
                cur_sub = name_cn
            continue

        try:
            mg = float(weight)
        except Exception:
            continue

        if name_cn in seen:
            continue
        seen.add(name_cn)

        entries.append(
            {
                "nameCn": name_cn,
                "nameLatin": name_latin,
                "wetWeightMg": mg,
                "taxonomy": {"group": _normalize_lvl1(cur_group or ""), "sub": cur_sub},
            },
        )

    return {"version": 1, "entries": entries}


def main() -> None:
    if not TAXONOMY_XLSX.exists():
        raise SystemExit(f"未找到文件：{TAXONOMY_XLSX}")
    if not WETWEIGHT_XLSX.exists():
        raise SystemExit(f"未找到文件：{WETWEIGHT_XLSX}")

    ASSETS_DIR.mkdir(parents=True, exist_ok=True)

    tax = parse_taxonomies(TAXONOMY_XLSX)
    wet = parse_wetweights(WETWEIGHT_XLSX)

    OUT_TAXONOMIES.write_text(json.dumps(tax, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    OUT_WETWEIGHTS.write_text(json.dumps(wet, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print("已生成：")
    print("-", OUT_TAXONOMIES)
    print("-", OUT_WETWEIGHTS)


if __name__ == "__main__":
    main()

